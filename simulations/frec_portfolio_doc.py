"""
Frec Portfolio Documentation
==============================
Produces a multi-sheet Excel workbook documenting the Frec portfolio:
  - Sheet 1  : Full weight matrix  (Ticker × Quarter, sorted by latest weight)
  - Sheet 2  : Top-30 holdings per quarter  (ranked snapshot)
  - Sheet 3  : Sector breakdown per quarter
  - Sheet 4  : Weight changes between consecutive quarters  (movers)

Output: data/frec_portfolio.xlsx

Run from the project root:
    py simulations/frec_portfolio_doc.py
"""

from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

DATA_DIR = Path("data")

# ── Palette ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1C2B3A"
C_HEADER_FG  = "F0F6FC"
C_ACCENT     = "3FB950"          # green
C_ROW_ALT    = "F6F8FA"
C_BORDER     = "D0D7DE"
C_UP         = "D4EDDA"          # light green (weight gained)
C_DOWN       = "F8D7DA"          # light red   (weight lost)
C_NEUTRAL    = "FFFFFF"

def hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(bold=True) -> Font:
    return Font(name="Segoe UI", bold=bold, color=C_HEADER_FG, size=10)

def body_font(bold=False) -> Font:
    return Font(name="Segoe UI", bold=bold, size=10)

def pct_fmt() -> str:
    return '0.000"%"'

# ── Load data ─────────────────────────────────────────────────────────────────
weight_dir = DATA_DIR / "sp500_weights"
quarters: dict[str, pd.Series] = {}
for f in sorted(weight_dir.glob("*.csv")):
    df = pd.read_csv(f)
    df["ticker"] = (df["ticker"].str.replace("/", "-", regex=False)
                                .str.replace(".", "-", regex=False))
    df = df.groupby("ticker")["weight_%"].sum()
    quarters[f.stem] = df

# Wide matrix: tickers as rows, quarters as columns
wide = pd.DataFrame(quarters).fillna(0.0)
wide.index.name = "Ticker"

# Sort by latest quarter weight descending
latest_q = sorted(quarters.keys())[-1]
wide = wide.sort_values(latest_q, ascending=False)
quarter_dates = sorted(quarters.keys())

# Sector info
constituents = pd.read_csv(DATA_DIR / "constituents.csv", index_col=0)
sector_map   = constituents["sector"].to_dict()
status_map   = constituents["status"].to_dict()

# ── Create workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 : Full weight matrix
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Full Weight Matrix")

# Title
ws1.merge_cells("A1:K1")
title_cell = ws1["A1"]
title_cell.value = "Frec Portfolio  —  Quarterly Weight Matrix (%)"
title_cell.font  = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
title_cell.fill  = hdr_fill(C_HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 26

ws1.merge_cells("A2:K2")
sub = ws1["A2"]
sub.value = (f"Source: SEC EDGAR NPORT-P filings (SPY / CIK 0000884394)  |  "
             f"{len(wide)} tickers  |  {len(quarter_dates)} quarters  |  "
             f"Weights normalised to equity-only universe (100%)")
sub.font      = Font(name="Segoe UI", italic=True, size=9, color="555555")
sub.fill      = hdr_fill("F0F6FC")
sub.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[2].height = 16

# Column headers: row 3
FIXED_COLS = ["Ticker", "Sector", "Status"]
all_cols   = FIXED_COLS + quarter_dates

for col_idx, col_name in enumerate(all_cols, start=1):
    cell = ws1.cell(row=3, column=col_idx, value=col_name)
    cell.font      = hdr_font()
    cell.fill      = hdr_fill(C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

ws1.row_dimensions[3].height = 32

# Data rows
for row_idx, ticker in enumerate(wide.index, start=4):
    sector = sector_map.get(ticker, "Unknown")
    status = status_map.get(ticker, "")
    row_bg = C_ROW_ALT if (row_idx % 2 == 0) else C_NEUTRAL
    fill   = PatternFill("solid", fgColor=row_bg)

    for col_idx, col_name in enumerate(all_cols, start=1):
        if col_name == "Ticker":
            val = ticker
        elif col_name == "Sector":
            val = sector
        elif col_name == "Status":
            val = status
        else:
            val = float(wide.loc[ticker, col_name])

        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = body_font(bold=(col_name == "Ticker"))
        cell.fill      = fill
        cell.border    = thin_border()
        cell.alignment = Alignment(horizontal="right" if col_name not in FIXED_COLS
                                   else "left", vertical="center")
        if col_name not in FIXED_COLS:
            cell.number_format = '0.000'

# Column widths
ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 26
ws1.column_dimensions["C"].width = 10
for i, _ in enumerate(quarter_dates, start=4):
    ws1.column_dimensions[get_column_letter(i)].width = 12

# Conditional formatting: colour scale on weight columns
weight_start = get_column_letter(len(FIXED_COLS) + 1)
weight_end   = get_column_letter(len(all_cols))
data_rows    = f"{weight_start}4:{weight_end}{3 + len(wide)}"
ws1.conditional_formatting.add(
    data_rows,
    ColorScaleRule(
        start_type="num",  start_value=0,   start_color="FFFFFF",
        mid_type="num",    mid_value=1,     mid_color="C8E6C9",
        end_type="num",    end_value=10,    end_color="1B5E20",
    )
)

# Freeze header rows and ticker column
ws1.freeze_panes = "D4"

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 : Top-30 per quarter
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Top-30 per Quarter")

ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "Frec Portfolio  —  Top 30 Holdings per Quarter"
t.font  = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
t.fill  = hdr_fill(C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

cur_col = 1
for q in quarter_dates:
    ser = quarters[q].sort_values(ascending=False).head(30)

    # Quarter date header spanning 3 columns: Rank, Ticker, Weight
    ws2.merge_cells(
        start_row=2, end_row=2,
        start_column=cur_col, end_column=cur_col + 2
    )
    qh = ws2.cell(row=2, column=cur_col, value=q)
    qh.font      = hdr_font()
    qh.fill      = hdr_fill(C_ACCENT)
    qh.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 22

    for sub_col, label in zip(range(cur_col, cur_col + 3),
                               ["Rank", "Ticker", "Weight %"]):
        c = ws2.cell(row=3, column=sub_col, value=label)
        c.font      = hdr_font()
        c.fill      = hdr_fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()
    ws2.row_dimensions[3].height = 20

    for rank, (ticker, weight) in enumerate(ser.items(), start=1):
        row = 3 + rank
        row_bg = C_ROW_ALT if rank % 2 == 0 else C_NEUTRAL
        fill   = PatternFill("solid", fgColor=row_bg)

        for offset, val in enumerate([rank, ticker, round(weight, 3)]):
            c = ws2.cell(row=row, column=cur_col + offset, value=val)
            c.font      = body_font(bold=(offset == 1))
            c.fill      = fill
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="center" if offset != 1 else "left")
            if offset == 2:
                c.number_format = '0.000'

    for offset in range(3):
        col_letter = get_column_letter(cur_col + offset)
        ws2.column_dimensions[col_letter].width = [6, 10, 10][offset]

    cur_col += 4   # gap column between quarters

ws2.freeze_panes = "A4"

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 : Sector breakdown per quarter
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Sector Breakdown")

# Build sector × quarter weight table
sector_rows: dict[str, dict[str, float]] = {}
for q in quarter_dates:
    for ticker, weight in quarters[q].items():
        sector = sector_map.get(ticker, "Unknown")
        sector_rows.setdefault(sector, {})[q] = \
            sector_rows.get(sector, {}).get(q, 0.0) + weight

sector_df = pd.DataFrame(sector_rows).T.fillna(0.0)
sector_df = sector_df.reindex(columns=quarter_dates)
sector_df = sector_df.sort_values(latest_q, ascending=False)

ws3.merge_cells("A1:J1")
t = ws3["A1"]
t.value = "Frec Portfolio  —  Sector Weight Breakdown per Quarter (%)"
t.font  = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
t.fill  = hdr_fill(C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

for col_idx, label in enumerate(["Sector"] + quarter_dates, start=1):
    c = ws3.cell(row=2, column=col_idx, value=label)
    c.font      = hdr_font()
    c.fill      = hdr_fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
ws3.row_dimensions[2].height = 32

for row_idx, sector in enumerate(sector_df.index, start=3):
    row_bg = C_ROW_ALT if row_idx % 2 == 0 else C_NEUTRAL
    fill   = PatternFill("solid", fgColor=row_bg)
    c = ws3.cell(row=row_idx, column=1, value=sector)
    c.font = body_font(bold=True); c.fill = fill
    c.border = thin_border()
    c.alignment = Alignment(horizontal="left")
    ws3.column_dimensions["A"].width = 28
    for col_idx, q in enumerate(quarter_dates, start=2):
        val = round(float(sector_df.loc[sector, q]), 3)
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font(); cell.fill = fill
        cell.border = thin_border()
        cell.number_format = '0.000'
        cell.alignment = Alignment(horizontal="right")
        ws3.column_dimensions[get_column_letter(col_idx)].width = 12

# Total row
total_row = len(sector_df) + 3
ws3.cell(row=total_row, column=1, value="TOTAL").font = body_font(bold=True)
for col_idx, q in enumerate(quarter_dates, start=2):
    val = round(float(sector_df[q].sum()), 3)
    c = ws3.cell(row=total_row, column=col_idx, value=val)
    c.font = body_font(bold=True)
    c.fill = hdr_fill("E8F5E9")
    c.number_format = '0.000'
    c.alignment = Alignment(horizontal="right")
    c.border = thin_border()

ws3.freeze_panes = "B3"

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 : Quarter-over-quarter weight changes (top movers)
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("QoQ Weight Changes")

ws4.merge_cells("A1:F1")
t = ws4["A1"]
t.value = "Frec Portfolio  —  Quarter-over-Quarter Weight Changes (pp)"
t.font  = Font(name="Segoe UI", bold=True, size=13, color=C_HEADER_FG)
t.fill  = hdr_fill(C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

cur_row = 2
for i in range(1, len(quarter_dates)):
    q_prev, q_curr = quarter_dates[i-1], quarter_dates[i]
    diff = (wide[q_curr] - wide[q_prev]).sort_values()
    top_gain = diff.nlargest(15)
    top_loss = diff.nsmallest(15)

    # Period header
    ws4.merge_cells(
        start_row=cur_row, end_row=cur_row,
        start_column=1,    end_column=6
    )
    ph = ws4.cell(row=cur_row, column=1,
                  value=f"{q_prev}  ->  {q_curr}")
    ph.font      = Font(name="Segoe UI", bold=True, size=11, color=C_HEADER_FG)
    ph.fill      = hdr_fill(C_HEADER_BG)
    ph.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[cur_row].height = 22
    cur_row += 1

    # Sub-headers: Gainers | Losers
    for col_base, label in [(1, "Top Gainers (+pp)"), (4, "Top Losers (-pp)")]:
        ws4.merge_cells(
            start_row=cur_row, end_row=cur_row,
            start_column=col_base, end_column=col_base + 2
        )
        sh = ws4.cell(row=cur_row, column=col_base, value=label)
        sh.font      = hdr_font()
        sh.fill      = hdr_fill(C_ACCENT if "Gainer" in label else "C0392B")
        sh.alignment = Alignment(horizontal="center")
    cur_row += 1

    for col_base, label in [(1, "Ticker"), (2, "Sector"), (3, "Change pp"),
                             (4, "Ticker"), (5, "Sector"), (6, "Change pp")]:
        c = ws4.cell(row=cur_row, column=col_base if col_base <= 3
                     else col_base - 3 + 3, value=label)
        c = ws4.cell(row=cur_row, column=col_base, value=label)
        c.font = hdr_font(); c.fill = hdr_fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    cur_row += 1

    gainers = list(top_gain.items())[::-1]   # largest first
    losers  = list(top_loss.items())

    for rank in range(15):
        for col_base, items in [(1, gainers), (4, losers)]:
            if rank >= len(items):
                continue
            ticker, delta = items[rank]
            sector = sector_map.get(ticker, "")
            is_up  = delta >= 0
            bg     = C_UP if is_up else C_DOWN
            fill   = PatternFill("solid", fgColor=bg)
            for offset, val in enumerate([ticker, sector, round(delta, 3)]):
                c = ws4.cell(row=cur_row + rank, column=col_base + offset,
                             value=val)
                c.font   = body_font(bold=(offset == 0))
                c.fill   = fill
                c.border = thin_border()
                c.alignment = Alignment(
                    horizontal="right" if offset == 2 else "left")
                if offset == 2:
                    c.number_format = '+0.000;-0.000'
        ws4.row_dimensions[cur_row + rank].height = 16

    cur_row += 16   # 15 data rows + 1 blank gap

for col_letter, width in [("A",10),("B",26),("C",12),("D",10),("E",26),("F",12)]:
    ws4.column_dimensions[col_letter].width = width

ws4.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
out_path = DATA_DIR / "frec_portfolio.xlsx"
wb.save(str(out_path))
print(f"Saved -> {out_path}")
print(f"  Sheet 1 : Full Weight Matrix  ({len(wide)} tickers x {len(quarter_dates)} quarters)")
print(f"  Sheet 2 : Top-30 per Quarter")
print(f"  Sheet 3 : Sector Breakdown per Quarter")
print(f"  Sheet 4 : QoQ Weight Changes (top 15 gainers/losers)")

# ── Console preview ───────────────────────────────────────────────────────────
print("\n" + "=" * 72)
print("  Frec Portfolio  |  Top 20 Holdings  |  Latest Quarter:", latest_q)
print("=" * 72)
print(f"  {'Rank':<5} {'Ticker':<8} {'Sector':<30}", end="")
for q in quarter_dates:
    print(f"  {q}", end="")
print()
print("  " + "-" * (43 + len(quarter_dates) * 13))

top20 = wide.head(20)
for rank, (ticker, row) in enumerate(top20.iterrows(), start=1):
    sector = sector_map.get(ticker, "Unknown")[:28]
    print(f"  {rank:<5} {ticker:<8} {sector:<30}", end="")
    for q in quarter_dates:
        print(f"  {row[q]:>9.3f}%  ", end="")
    print()

print("\n  Sector totals per quarter:")
print(f"  {'Sector':<30}", end="")
for q in quarter_dates:
    print(f"  {q}", end="")
print()
for sector in sector_df.index:
    if sector == "Unknown":
        continue
    print(f"  {sector:<30}", end="")
    for q in quarter_dates:
        print(f"  {sector_df.loc[sector,q]:>9.3f}%  ", end="")
    print()
print(f"  {'TOTAL':<30}", end="")
for q in quarter_dates:
    print(f"  {sector_df[q].sum():>9.3f}%  ", end="")
print()
